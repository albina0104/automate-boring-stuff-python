# Spreadsheet to Text Files
# Opens a spreadsheet and writes the cells of column A into one text file,
# the cells of column B into another text file, and so on.

import os
import sys
import openpyxl
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def spreadsheet_to_text_files(directory, excel_file, basename):
    try:
        os.chdir(directory)
    except Exception as e:
        logging.error(f'Cannot change directory to {directory}: {e}')
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        logging.error(f'Cannot open the Excel file {excel_file}: {e}')
        sys.exit(1)

    sheet = wb.active

    txt_file_idx = 0
    for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=sheet.max_row):
        txt_file_idx += 1
        txt_filename = f'{basename}{txt_file_idx}.txt'
        with open(txt_filename, 'w') as txt_file:
            logging.debug(f'Column {txt_file_idx} will be written to the file {txt_filename}')
            txt_file.writelines(
                ('' if cell.value is None else str(cell.value)) + '\n' for cell in col
            )


if __name__ == '__main__':
    logging.debug('Start of the program')
    directory = 'text_files'
    excel_file = 'example.xlsx'
    basename = 'myFile'
    spreadsheet_to_text_files(directory, excel_file, basename)
    logging.debug('End of the program')
