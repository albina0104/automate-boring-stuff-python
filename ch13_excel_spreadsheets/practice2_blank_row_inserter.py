#! /usr/bin/env python3
# blank_row_inserter.py
# It takes two integers and a filename string as command line arguments.
# Let’s call the first integer N and the second integer M.
# Starting at row N, the program should insert M blank rows into the spreadsheet.
# Example how to run it:
# python blank_row_inserter.py 3 2 myProduce.xlsx
# Note: the solution is not suitable for spreadsheets with formulas, the formulas will NOT be adjusted accordingly!

import sys
import openpyxl
import logging
import traceback

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def insert_blank_rows(N, M, filename):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                new_row = cell.row + M if cell.row >= N else cell.row    # cell.row returns the row number
                new_sheet.cell(row=new_row, column=cell.column, value=cell.value)

        new_wb.save('insertedBlankRows.xlsx')
        print('Done!')
    except FileNotFoundError:
        print('File not found. Please provide a valid filename.')
    except Exception as e:
        print(f'An error occurred: {e}')
        logging.debug(f'Error: {e}. {traceback.format_exc()}')


if __name__ == "__main__":
    try:
        first_blank_row = int(sys.argv[1])
        num_blank_rows = int(sys.argv[2])
        filename = sys.argv[3]
        insert_blank_rows(first_blank_row, num_blank_rows, filename)
    except ValueError:
        print('You need to run the program with correct arguments: row number, number of rows, filename')
    except IndexError:
        print('Please provide all necessary arguments: row number, number of rows, filename')
    except Exception as e:
        print(f'An error occurred: {e}')
