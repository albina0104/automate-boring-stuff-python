# Spreadsheet Cell Inverter
# Inverts the row and column of the cells in the spreadsheet.
# For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
# This is done for all cells in the spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

sheet_data = []

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    sheet_data.append([])
    for cell in row:
        sheet_data[cell.row - 1].append(cell.value)

for y in range(len(sheet_data)):
    for x in range(len(sheet_data[y])):
        new_sheet.cell(row=x+1, column=y+1, value=sheet_data[y][x])

new_wb.save('invertedCells.xlsx')
print('Done!')
