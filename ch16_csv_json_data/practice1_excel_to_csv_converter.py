# Excel-to-CSV Converter
# Reads all the Excel files in the current working directory and outputs them as CSV files.

import openpyxl, csv, os
from pathlib import Path

for excel_filename in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not excel_filename.endswith('.xlsx'):
        continue
    workbook = openpyxl.load_workbook(excel_filename)

    for sheet_name in workbook.sheetnames:
        # Loop through every sheet in the workbook.
        sheet = workbook[sheet_name]

        # Create the CSV filename from the Excel filename and sheet title.
        csv_filename = Path(excel_filename).stem + '_' + sheet_name + '.csv'
        # Create the csv.writer object for this CSV file.
        csv_file = open(csv_filename, 'w', newline='')
        csv_writer = csv.writer(csv_file)

        # Loop through every row in the sheet.
        for row_num in range(1, sheet.max_row + 1):
            row_data = []    # append each cell to this list
            # Loop through each cell in the row.
            for col_num in range(1, sheet.max_column + 1):
                # Append each cell's data to row_data.
                row_data.append(sheet.cell(row=row_num, column=col_num).value)

            # Write the row_data list to the CSV file.
            csv_writer.writerow(row_data)

        csv_file.close()

print('Done!')
